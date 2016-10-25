# Movie Assistor - This script scans the directory for movies and returns the ratings and riviews in a spreadsheet.

import os
import guessit
import threading
import pyperclip
import sys
import time
import pprint
import getInfo
import subprocess
import openpyxl
from openpyxl.styles import Font
# List of all possible video formats
name_list = ["webm","mkv","flv","vob","ogv","ogg","drc","gif","gifv","mng","avi","moc","qt","wmv","yuv","rm","rmvb","asf",
			 "mp4","m4p","m4v","mpg","mp2","mpe","mpeg","m2v","svi","3gp","3g2","mxf","roq","nsv","f4p","f4a","f4b"]

#list will contain movie name_list
movie_names = []

f_list = []

#list will contain movie information
movie_info = ["","","","","","","",""]

# path of directory is stored here
dir_path = ""

# Input of directory path
def input_directory():

	if len(sys.argv) == 1:
		print("Copy folder's path and paste here : ")
		dir_path = input()
		#dir_path = "/home/kiit/movies"
	else:
		dir_path = pyperclip.paste()

	if os.path.exists(dir_path):
		os.chdir(dir_path)

	else:
		print("Directory Does Not Exists")
		return 999

# movie name extracting
def movie_title(item):
	title = ""
	if "sample" in item.lower():
		return ""
	try:
		data = guessit.guessit(item)
	except:
		return ""

	try:
		extension = data['container']
	except:
		return ""

	if extension in name_list:
		try:
			title = data['title']
		except:
			title = ""
	return title

# Walkthrough all movie folders
def walk_dir():
	title=""

	for	folderName,	subfolders,	filenames in os.walk(os.getcwd()):
		for	filename in	filenames:
			title = movie_title(filename)
			if title != "":
				movie_names.append(title)

#finally calling and displaying for testing purpose
def func_call():
    input_directory()
    walk_dir()
    #print(len(movie_names)
    download_review()

def thread_download(num):
	#download information
	temp_list = list()
	temp = num
	while True:
		try:
			check = getInfo.getInfo(movie_names[num])
			temp_list.append(check)
			if check == "":
				movie_info[temp] = temp_list
						
				return movie_info[temp]
			else:
				num = num + 8
		except:
			movie_info[temp] = temp_list
			return movie_info[temp]


def download_review():
	thread_list = list()
	for i in range(0,8):
		temp_thread = threading.Thread(target = thread_download, args=[i])
		thread_list.append(temp_thread)
		temp_thread.start()
	for threadObj in thread_list:
		threadObj.join()
	
	

def display():
	
	for i in f_list:
		pprint.pprint(i)
		

def saveToSpreadSheet():
	print("Writing to spreadsheet....")
	wb = openpyxl.Workbook()
	wb.create_sheet(index=0,title="My Movie Database")
	sheet = wb.active
	sheet['A1'].font = Font(size=12,bold=True)
	sheet['B1'].font = Font(size=12,bold=True)
	sheet['C1'].font = Font(size=12,bold=True)
	sheet['D1'].font = Font(size=12,bold=True)
	sheet['E1'].font = Font(size=12,bold=True)
	sheet['F1'].font = Font(size=12,bold=True)
	sheet['G1'].font = Font(size=12,bold=True)
	sheet['H1'].font = Font(size=12,bold=True)
	sheet['I1'].font = Font(size=12,bold=True)
	
	sheet.column_dimensions['A'].width = 22
	sheet.column_dimensions['C'].width = 12
	sheet.column_dimensions['D'].width = 15
	sheet.column_dimensions['E'].width = 20
	sheet.column_dimensions['G'].width = 30
	sheet.column_dimensions['H'].width = 30
	sheet.column_dimensions['I'].width = 70

	sheet['A1']="Title"
	sheet['B1']="Year"
	sheet['C1']="Genre"
	sheet['D1']="IMDB"
	sheet['E1']="Rotten Tomato"
	sheet['F1']="Runtime"
	sheet['G1']="Actors"
	sheet['H1']="Plot"
	sheet['I1']="Director"
	k=2
	f_list=movie_info[0]+movie_info[1]+movie_info[2]+movie_info[3]+movie_info[4]+movie_info[5]+movie_info[6]+movie_info[7]
	movie_sorted = sorted(f_list,key=lambda x:x["imdbRating"],reverse=True)
	#print(f_list)	
	for info in f_list:
		#print(info)
		if info['Title'] == 'N/A':
			continue
		sheet.cell(row=k,column=1).value = str(info["Title"])
		sheet.cell(row=k,column=2).value = str(info["Year"])
		sheet.cell(row=k,column=3).value = str(info["Genre"])
		sheet.cell(row=k,column=4).value = str(info["imdbRating"])
		sheet.cell(row=k,column=5).value = str(info["tomatoRating"])
		sheet.cell(row=k,column=6).value = str(info["Runtime"])
		sheet.cell(row=k,column=7).value = str(info["Actors"])
		sheet.cell(row=k,column=8).value = str(info["Plot"])
		sheet.cell(row=k,column=9).value = str(info["Director"])
		k=k+1
	wb.save("My Movie Database.xlsx")
	time.sleep(2)

def openFile():
	subprocess.Popen(['see','My Movie Database.xlsx'])

func_call()
#display()
saveToSpreadSheet()
openFile()
