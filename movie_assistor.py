# Movie Assistor - This script scans the directory for movies and returns the ratings and riviews in a spreadsheet.
import os
import guessit
import threading
import pyperclip
import sys
import time
import getInfo
import openpyxl
import subprocess
from openpyxl.styles import	Font,Style
# List of all possible video formats
name_list = ["webm","mkv","flv","vob","ogv","ogg","drc","gif","gifv","mng","avi","moc","qt","wmv","yuv","rm","rmvb","asf",
			 "mp4","m4p","m4v","mpg","mp2","mpe","mpeg","m2v","svi","3gp","3g2","mxf","roq","nsv","f4p","f4a","f4b"]

#list will contain movie name_list
movie_names = []

#list will contain movie information
movie_info = []

# path of directory is stored here
dir_path = ""

# Input of directory path
def input_directory():

	if len(sys.argv) == 1:
		print("Copy folder's path and paste here : ")
		dir_path = input()
	else:
		dir_path = pyperclip.paste()


	if os.path.exists(dir_path):
		os.chdir(dir_path)

	else:
		print("Directory Does Not Exists")
		return 999

# movie name extracting
def movie_title(item):
	title = ""
	if "sample" in item.lower():
		return ""
	try:
		data = guessit.guessit(item)
	except:
		return ""

	try:
		extension = data['container']
	except:
		return ""

	if extension in name_list:
		try:
			title = data['title']
		except:
			title = ""
	return title

# Walkthrough all movie folders
def walk_dir():
	title=""

	for	folderName,	subfolders,	filenames in os.walk(os.getcwd()):
		for	filename in	filenames:
			title = movie_title(filename)
			if title != "":
				movie_names.append(title)
	#movie_names=list(set(movie_names))

#finally calling and displaying for testing purpose
def func_call():
    input_directory()
	#print(dir_path)
    walk_dir()

    for name in movie_names:
        print("Getting Info of "+name+"...")
        downFile(name)


def display():
	for i in movie_info:
		print(i)

def downFile(mname):
    x=getInfo.getInfo(mname)
    #print(x)
    movie_info.append(x)

def saveToSpreadSheet():
	print("Writing to spreadsheet....")
	wb = openpyxl.Workbook()
	wb.create_sheet(index=0,title="My Movie Database")
	sheet = wb.active
	sheet['A1'].font = Font(size=12,bold=True)
	sheet['B1'].font = Font(size=12,bold=True)
	sheet['C1'].font = Font(size=12,bold=True)
	sheet['D1'].font = Font(size=12,bold=True)
	sheet['E1'].font = Font(size=12,bold=True)
	sheet['F1'].font = Font(size=12,bold=True)
	sheet['G1'].font = Font(size=12,bold=True)
	sheet['H1'].font = Font(size=12,bold=True)
	sheet['I1'].font = Font(size=12,bold=True)
	#italic = openpyxl.styles.Font(size=28,bold=True)
	#styleobj= openpyxl.styles.Style(font=italic24font)
	#sheet['A'].style/styleobj
	sheet.column_dimensions['A'].width = 22
	sheet.column_dimensions['C'].width = 12
	sheet.column_dimensions['D'].width = 15
	sheet.column_dimensions['E'].width = 20
	sheet.column_dimensions['G'].width = 30
	sheet.column_dimensions['H'].width = 30
	sheet.column_dimensions['I'].width = 70
	#sheet.freeze_panes='A1'
	#sheet.freeze_panes='B1'
	#sheet.freeze_panes='C1'
	#sheet.freeze_panes='D1'
	#sheet.freeze_panes='E1'
	#sheet.freeze_panes='F1'
	#sheet.freeze_panes='G1'
	#sheet.freeze_panes='H1'
	#sheet.freeze_panes='I1'

	sheet['A1']="Title"
	sheet['B1']="Year"
	sheet['C1']="Genre"
	sheet['D1']="IMDB"
	sheet['E1']="Rotten Tomato"
	sheet['F1']="Runtime"
	sheet['G1']="Actors"
	sheet['H1']="Director"
	sheet['I1']="Plot"
	k=2
	for info in movie_info:
		if info['Title'] == 'N/A':
			continue
		sheet.cell(row=k,column=1).value = str(info["Title"])
		sheet.cell(row=k,column=2).value = str(info["Year"])
		sheet.cell(row=k,column=3).value = str(info["Genre"])
		sheet.cell(row=k,column=4).value = str(info["imdbRating"])
		sheet.cell(row=k,column=5).value = str(info["tomatoRating"])
		sheet.cell(row=k,column=6).value = str(info["Runtime"])
		sheet.cell(row=k,column=7).value = str(info["Actors"])
		sheet.cell(row=k,column=8).value = str(info["Director"])
		sheet.cell(row=k,column=9).value = str(info["Plot"])
		k=k+1
	wb.save("My Movie Database.xlsx")
	time.sleep(2)

def openFile():
	subprocess.Popen(['see','My Movie Database.xlsx'])

func_call()
saveToSpreadSheet()
openFile()
